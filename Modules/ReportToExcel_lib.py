import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series)
from Modules import Date
#import Date

cb = Border(left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin'))

def export(report, filename):
    wb = openpyxl.Workbook()
    wb.active.title = "Report"
    sheet = wb["Report"]

    report["DATE"] = Date.toStrMDY(report["DATE"])
    report["REF"]["EXP_DATE"] = Date.toStrMDY(report["REF"]["EXP_DATE"])
    report["DMM"]["EXP_DATE"] = Date.toStrMDY(report["DMM"]["EXP_DATE"])

    ### HEADER
    Type = report["TYPE"] #<"STD", "CAL"
    col = "A"
    row = 1
    dat = report
    row = expandHeader(sheet, dat, col, row)

    ### REF
    col = "A"
    row += 1
    key = "REF"
    dat = report[key]
    row = expandDev(sheet, dat, col, row, key)

    ### DMM
    col = "A"
    row += 1
    key = "DMM"
    dat = report[key]
    row = expandDev(sheet, dat, col, row, key)

    ### START ITS COEF
    col = "A"
    row += 1
    key = "COEF"
    dat = report[key]
    row = expandITS(sheet, dat, col, row, "START COEF")

    ### RESULTS - HEADER
    col = "A"
    row += 1
    key = "RESULTS"
    dat = report[key]
    row = expandResultsHdr(sheet, dat, col, row, key)

    if Type == "STD":
        ### RESULTS - NEW COEF
        col = "B"
        #row += 1
        key = "NEW_COEF"
        dat = report["RESULTS"][key]
        row = expandITS(sheet, dat, col, row, "NEW COEF")
    
    elif Type == "CAL":
        ...
    
    ### RESULTS - DEVIATION TABLE
    col = "A"
    row += 1
    key = "DEVTBL"
    dat = report["RESULTS"][key]
    tblRow = row
    tblCol = chr(ord(col) +1)
    row = expandDevTbl(sheet, dat, col, row, key)

    ### RESULTS - DEVIATION PLOT
    col = "A"
    row += 1
    units = report["RESULTS"]["UNITS"]
    addDevPlot(sheet, dat, (col, row), tblCol, tblRow, units=units)

    ### RESULTS - TEST PROFILE PLOT
    col = "A"
    row += 15
    units = report["RESULTS"]["UNITS"]
    addTempCurvePlot(sheet, dat, (col, row), tblCol, tblRow, units=units)

    if (filename.split('.'))[-1] != "xlsx":
        filename = "%s.xlsx"%(filename)
    wb.save( filename )
    return wb 

def expandHeader(sheet, dat, col, row) -> int:
    hdrs = ("Test type","Unit ID","DATE (M/D/Y)")
    keys = ("TYPE","UID","DATE")
    return _expand(sheet, dat, col, row, keys, keys)
    
def expandDev(sheet, dat, col, row, title) -> int:
    """  """
    ## Title
    cell = "%c%d"%(col,row)
    sheet[cell] = title
    sheet[cell].font = Font(bold=True)
    sheet[cell].border = cb
    col = chr(ord(col)+1)

    ## Info
    hdrs = ("MFR","MODEL","SN","EXP_DATE")
    keys = ("MFR","MODEL","SN","EXP_DATE")
    return _expand(sheet, dat, col, row, keys, keys)

def expandITS(sheet, dat, col, row, title) -> int:
    """  """
    ## Title
    cell = "%c%d"%(col,row)
    sheet[cell] = title
    sheet[cell].font = Font(bold=True)
    sheet[cell].border = cb
    col = chr(ord(col)+1)

    ## Info
    if dat["TYPE"] == "ITS90":
        keys = ("TYPE","RTPW","a4","b4","a7","b7","c7")
    else:
        keys = ("TYPE","R0","alpha","delta")

    return _expand(sheet, dat, col, row, keys, keys)

def expandResultsHdr(sheet, dat, col, row, title) -> int:
    ## Title
    cell = "%c%d"%(col,row)
    sheet[cell] = title
    sheet[cell].font = Font(bold=True)
    sheet[cell].border = cb
    col = chr(ord(col)+1)

    hdrs = ("UNITS","Ambient")
    keys = ("UNITS","AMB")
    return _expand(sheet, dat, col, row, hdrs, keys)

def expandDevTbl(sheet, dat, col, row, title) -> int:
    """  """
    ## Title
    cell = "%c%d"%(col,row)
    sheet[cell] = title
    sheet[cell].font = Font(bold=True)
    sheet[cell].border = cb
    col = chr(ord(col)+1)

    ## Info
    keys = dat[0].keys()
    hdrRow = row
    
    row = _expandLine(sheet, dat, col, row, keys, bold=True)
    datRow = row

    for line in dat:
        vals = []
        for key in keys:
            vals.append( line[key] )
        row = _expandLine(sheet, dat, col, row, vals, bold=True)
    return row
    
def addDevPlot(sheet, dat, pos, hdrCol, hdrRow, units='C'):
    chart = ScatterChart()
    chart.title = "Test points"
    chart.style = 13
    chart.x_axis.title = "Reference temperature (%s)"%(units)
    chart.y_axis.title = "Deviation from reference (%s)"%(units)
    chart.legend = None          #< No legend box

    xName = "REF"
    yName = "DEV"    
    xCol = None
    yCol = None
    if isinstance(hdrCol, str):
        col = ord(hdrCol) -ord("A")
        
    for col in range(col,col+15):
        if sheet.cell(hdrRow,col).value == xName:
            xCol = col
        if sheet.cell(hdrRow,col).value == yName:
            yCol = col

    if (xCol != None) and (yCol != None):
        stRow = hdrRow +1
        endRow = stRow +22
        xvalues = Reference(sheet, min_col=xCol, min_row=stRow, max_row=endRow)
        yvalues = Reference(sheet, min_col=yCol, min_row=stRow, max_row=endRow)
        series = Series(yvalues, xvalues, title_from_data=False)
        series.graphicalProperties.line.solidFill = "000000" #<RGB black
        series.graphicalProperties.line.width = 20050 #<EMUs
        chart.series.append(series)
        cell = "%c%d"%(pos[0], pos[1])
        sheet.add_chart(chart, cell)

def addTempCurvePlot(sheet, dat, pos, hdrCol, hdrRow, units='C'):
    chart = ScatterChart()
    chart.title = "Temperature profile"
    chart.style = 13
    chart.x_axis.title = "Time (Seconds)"
    chart.y_axis.title = "UUT temperature (%s)"%(units)
    #chart.legend.position = 'b' #< Legend box at the bottom
    chart.legend = None          #< No legend box

    xName = "TIME"
    _yName = ("THRS_MIN", "THRS_MAX", "UUT")
    _colors = ("000000", "FF0000", "FF0000") #< Limits in read and curve in black
    xCol = None
    yCol = None
    if isinstance(hdrCol, str):
        col = ord(hdrCol) -ord("A")

    _yCol = []
    for col in range(col,col+15):
        if sheet.cell(hdrRow,col).value == xName:
            xCol = col

        for yName in _yName:
            if sheet.cell(hdrRow,col).value == yName:
                _yCol.append(col)

    if (xCol != None) and (len(_yCol) > 0):
        for yCol, color in zip(_yCol, _colors):
            stRow = hdrRow +1
            endRow = stRow +22
            xvalues = Reference(sheet, min_col=xCol, min_row=stRow, max_row=endRow)
            yvalues = Reference(sheet, min_col=yCol, min_row=stRow, max_row=endRow)
            series = Series(yvalues, xvalues, title_from_data=False)
            series.graphicalProperties.line.solidFill = color #<RGB
            series.graphicalProperties.line.width = 20050 #<EMUs
            chart.series.append(series)

        cell = "%c%d"%(pos[0], pos[1])
        sheet.add_chart(chart, cell)

def _expandLine(sheet, dat, col, row, vals, bold=False):
    for val in vals:
        cell = "%c%d"%(col,row)
        sheet[cell] = val
        sheet[cell].font = Font(bold=bold)
        sheet[cell].border = cb
        col = chr(ord(col)+1)
    return row +1

def _expand(sheet, dat, col, row, hdrs, keys):
    for hdr, key in zip(hdrs, keys):
        cell = "%c%d"%(col,row)
        sheet[cell] = hdr
        sheet[cell].font = Font(bold=True)
        sheet[cell].border = cb

        cell = "%c%d"%( chr(ord(col)+1),row)
        if not key in dat:
            key = key[0].upper() +key[1:]
            if not key in dat:
                key = key.upper()

            if key == "SN":
                key = "UID"
        sheet[cell] = dat[key]
        sheet[cell].border = cb
        row += 1
    return row



if __name__ == "__main__":
    if True:
        report = {"TYPE": "STD",
                  "UID": "A1001",
                  "DATE": {"DAY": 21, "MONTH": 4, "YEAR": 2022},
                  "UUT": {"MFR": "KNC", "MODEL": "3605A", "UID": "A1001",
                          "EXP_DATE": {"YEAR": 2022, "MONTH": 6, "DAY": 19},
                          "COEF": {"TYPE": "IEEE488", "ADDR": "GPIB0::1::INSTR"}},
                  
                  "DMM": {"MFR": "HP", "MODEL": "HP34401A", "SN": "4886",
                          "EXP_DATE": {"YEAR": 22, "MONTH": 5, "DAY": 19},
                          "COEF": {"TYPE": "IEEE488", "LANG": "Fluke", "ADDR": "GPIB0::6::INSTR"}},
                  
                  "REF": {"MFR": "Fluke", "MODEL": "5628", "SN": "2011",
                          "EXP_DATE": {"YEAR": 22, "MONTH": 5, "DAY": 19},
                          "COEF": {"TYPE": "ITS90",
                                   "RTPW": 25.49449,
                                   "a4": 5.302158e-05,
                                   "b4": 2.200709e-07,
                                   "a7": -1.052826e-05,
                                   "b7": 4.104768e-05,
                                   "c7": -2.272356e-05}},
                  
                  "COEF": {"TYPE": "ITS27",
                           "R0": 199.324463,
                           "Alpha": 0.004025,
                           "Delta": 0.612724},
                  
                  "RESULTS": {"UNITS": "C",
                              "AMB": 22.147000000000002,
                              "DEVTBL": [{"UNIT": -25.001, "RES": 179.098, "REF": -25.17, "DEV": 0.16, "UNITS": "C", "AMB": 23.85, "SOAK": 5.0},
                                         {"UNIT": 39.999, "RES": 231.53, "REF": 39.75, "DEV": 0.23, "UNITS": "C", "AMB": 20.77, "SOAK": 5.0},
                                         {"UNIT": 104.999, "RES": 283.526, "REF": 105.73, "DEV": -0.73, "UNITS": "C", "AMB": 21.81, "SOAK": 5.0}],
                              
                              "NEW_COEF": {"TYPE": "ITS27",
                                           "R0": 199.57196,
                                           "Alpha": 0.003983,
                                           "Delta": 1.859594}}}
    else:
        report = {"TYPE": "CAL",
                  "UID": "ENG (6-1)",
                  "DATE": {"DAY": 28, "MONTH": 4, "YEAR": 2022},
                  "UUT": {"MFR": "KNC", "MODEL": "3604U", "UID": "ENG (6-1)",
                          "EXP_DATE": {"YEAR": 2022, "MONTH": 6, "DAY": 19},
                          "COEF": {"TYPE": "IEEE488", "ADDR": "GPIB0::1::INSTR"}},
                  
                  "DMM": {"MFR": "Agilent", "MODEL": "HP34401A", "SN": "MY45037939",
                          "EXP_DATE": {"YEAR": 2022, "MONTH": 5, "DAY": 19},
                          "COEF": {"TYPE": "IEEE488", "LANG": "Fluke", "ADDR": "GPIB0::10::INSTR"}},
                  
                  "REF": {"MFR": "Rosemount", "MODEL": "162CE", "SN": "1666",
                          "EXP_DATE": {"YEAR": 2025, "MONTH": 2, "DAY": 10},
                          "COEF": {"TYPE": "ITS90",
                                   "RTPW": 25.60376,
                                   "a4": -0.000103365082,
                                   "b4": -5.81762439e-06,
                                   "a7": -9.48187167e-05,
                                   "b7": -2.72062893e-05,
                                   "c7": 6.570505e-06}},
                  
                  "COEF": {"TYPE": "ITS27",
                           "R0": 201.66746521,
                           "Alpha": 0.00391811,
                           "Delta": 1.48850858,
                           "AMB": 29.6},
                  
                  "RESULTS": {"UNITS": "C",
                              "AMB": 27.9525,
                              "DEVTBL": [{"UNIT":  37.78, "REF":  37.78, "UUT":  37.78, "UNITS": "C", "DEV":-0.0068, "AMB": 26.5, "SOAK": 25.0, "MARK": "PASS"},
                                         {"UNIT":  93.34, "REF":  93.30, "UUT":  93.34, "UNITS": "C", "DEV": 0.0449, "AMB": 26.9, "SOAK": 25.0, "MARK": "PASS"},
                                         {"UNIT": 148.89, "REF": 148.83, "UUT": 148.89, "UNITS": "C", "DEV": 0.0600, "AMB": 26.0, "SOAK": 25.0, "MARK": "PASS"},
                                         {"UNIT": 204.44, "REF": 204.37, "UUT": 204.44, "UNITS": "C", "DEV": 0.0654, "AMB": 26.8, "SOAK": 25.0, "MARK": "PASS"},
                                         {"UNIT": 260.00, "REF": 259.89, "UUT": 260.00, "UNITS": "C", "DEV": 0.1005, "AMB": 27.6, "SOAK": 25.0, "MARK": "PASS"},
                                         {"UNIT": 315.56, "REF": 315.43, "UUT": 315.56, "UNITS": "C", "DEV": 0.1224, "AMB": 27.0, "SOAK": 25.0, "MARK": "PASS"},
                                         {"UNIT": 371.11, "REF": 370.98, "UUT": 371.11, "UNITS": "C", "DEV": 0.1294, "AMB": 27.7, "SOAK": 25.0, "MARK": "PASS"},
                                         {"UNIT": 426.67, "REF": 426.56, "UUT": 426.67, "UNITS": "C", "DEV": 0.1072, "AMB": 28.7, "SOAK": 25.0, "MARK": "PASS"},
                                         {"UNIT": 482.22, "REF": 482.10, "UUT": 482.22, "UNITS": "C", "DEV": 0.1153, "AMB": 28.4, "SOAK": 25.0, "MARK": "PASS"},
                                         {"UNIT": 537.78, "REF": 537.68, "UUT": 537.78, "UNITS": "C", "DEV": 0.0985, "AMB": 29.8, "SOAK": 25.0, "MARK": "PASS"},
                                         {"UNIT": 593.33, "REF": 593.21, "UUT": 593.33, "UNITS": "C", "DEV": 0.1230, "AMB": 29.9, "SOAK": 25.0, "MARK": "PASS"},
                                         {"UNIT": 621.11, "REF": 620.96, "UUT": 621.11, "UNITS": "C", "DEV": 0.1463, "AMB": 30.0, "SOAK": 25.0, "MARK": "PASS"}]}}

    export(report, "test")
